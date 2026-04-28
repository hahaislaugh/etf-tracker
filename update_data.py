#!/usr/bin/env python3
"""
ETF 持股資料自動更新腳本
每日台北時間晚間 11:00 執行（本機透過 launchd，雲端透過 GitHub Actions）

支援 ETF：
  00991A  復華台灣未來50 —— GET https://www.fhtrust.com.tw/api/assetsExcel/ETF23/{YYYYMMDD}
  00992A  群益科技創新   —— POST https://www.capitalfund.com.tw/CFWeb/api/etf/buyback
  00981A  統一台股增長   —— GET https://www.ezmoney.com.tw/ETF/Fund/Info?fundCode=49YTW

使用方式：
  python3 update_data.py            # 下載今天的資料並重新產生 HTML
  python3 update_data.py --dry-run  # 只解析現有檔案，不下載
"""

import os, sys, glob, json, datetime, urllib.request, urllib.error, http.cookiejar
import openpyxl
from zoneinfo import ZoneInfo

SCRIPT_DIR    = os.path.dirname(os.path.abspath(__file__))
FUHUA_DIR     = os.path.join(SCRIPT_DIR, "復華")
QUNYI_DIR     = os.path.join(SCRIPT_DIR, "群益")
QUNYI982_DIR  = os.path.join(SCRIPT_DIR, "群益982")
UNITRUST_DIR  = os.path.join(SCRIPT_DIR, "統一")
NOMURA_DIR    = os.path.join(SCRIPT_DIR, "野村")
HTML_OUT      = os.path.join(SCRIPT_DIR, "index.html")
TAIPEI_TZ     = ZoneInfo("Asia/Taipei")

FUHUA_API     = "https://www.fhtrust.com.tw/api/assetsExcel/ETF23/{date}"
QUNYI_API     = "https://www.capitalfund.com.tw/CFWeb/api/etf/buyback"
UNITRUST_URL  = "https://www.ezmoney.com.tw/ETF/Fund/Info?fundCode=49YTW"
NOMURA_API    = "https://www.nomurafunds.com.tw/API/ETFAPI/api/Fund/GetFundAssets"

for d in (FUHUA_DIR, QUNYI_DIR, QUNYI982_DIR, UNITRUST_DIR, NOMURA_DIR):
    os.makedirs(d, exist_ok=True)


def taipei_today():
    return datetime.datetime.now(TAIPEI_TZ).date()


def is_weekday(d: datetime.date) -> bool:
    return d.weekday() < 5  # Mon-Fri


def download_file(url: str, dest: str) -> bool:
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=15) as r:
            data = r.read()
        if data[:2] != b"PK":
            print(f"  ⚠️  回應不是 xlsx（可能是 HTML 錯誤頁），跳過：{url}")
            return False
        with open(dest, "wb") as f:
            f.write(data)
        print(f"  ✓ 已下載 → {dest}")
        return True
    except urllib.error.HTTPError as e:
        print(f"  ✗ HTTP {e.code}：{url}")
        return False
    except Exception as e:
        print(f"  ✗ 錯誤：{e}")
        return False


def fetch_qunyi_api(date_str: str = None, fund_id: str = "500") -> list:
    """透過群益投信 JSON API 取得完整持股清單。
    date_str: 'YYYY/MM/DD' 格式；None 表示最新日期。
    fund_id: '500'=00992A, '399'=00982A
    """
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
        "Accept": "application/json, text/plain, */*",
        "Content-Type": "application/json",
        "Referer": f"https://www.capitalfund.com.tw/etf/product/detail/{fund_id}/portfolio",
        "Origin": "https://www.capitalfund.com.tw",
    }
    body = json.dumps({"fundId": fund_id, "date": date_str}).encode()
    req = urllib.request.Request(QUNYI_API, data=body, headers=headers, method="POST")
    with urllib.request.urlopen(req, timeout=15) as r:
        j = json.loads(r.read())
    if j.get("code") != 200:
        raise ValueError(f"群益 API 回傳 code={j.get('code')}")
    raw = j["data"]["stocks"]
    stocks = [{
        "code":   s["stocNo"],
        "name":   s["stocName"],
        "shares": int(s["share"]),
        "weight": round(float(s["weightRound"]), 4),
    } for s in raw]
    return stocks


def fetch_nomura_api(date_str: str = None) -> tuple:
    """從野村投信 API 取得 00980A 持股資料。回傳 (date_str, stocks)。
    date_str: 'YYYY/MM/DD'；None 表示最新日期。
    """
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
        "Content-Type": "application/json",
        "Referer": "https://www.nomurafunds.com.tw/ETFWEB/product-description?fundNo=00980A&tab=Shareholding",
        "Origin": "https://www.nomurafunds.com.tw",
    }
    # API 要求 YYYY-MM-DD 格式；None 表示最新
    api_date = date_str.replace("/", "-") if date_str else None
    body = json.dumps({"FundID": "00980A", "SearchDate": api_date}).encode()
    req = urllib.request.Request(NOMURA_API, data=body, headers=headers, method="POST")
    with urllib.request.urlopen(req, timeout=15) as r:
        j = json.loads(r.read())
    if j.get("StatusCode") != 0:
        raise ValueError(f"野村 API 回傳 StatusCode={j.get('StatusCode')}")
    data = j["Entries"]["Data"]
    nav_date = data["FundAsset"]["NavDate"]  # "YYYY/MM/DD"
    rows = data["Table"][0]["Rows"]
    if not rows:
        raise ValueError("野村 API 回傳空資料")
    stocks = [{
        "code":   str(r[0]).strip(),
        "name":   str(r[1]).strip(),
        "shares": int(str(r[2]).replace(",", "")),
        "weight": round(float(r[3]), 4),
    } for r in rows if r[0]]
    return nav_date, stocks


def fetch_unitrust_ezmoney() -> tuple:
    """從 ezmoney 頁面擷取 00981A 持股資料。回傳 (date_str, stocks)。
    date_str 格式: 'YYYY/MM/DD'
    """
    cj = http.cookiejar.CookieJar()
    opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj))
    req = urllib.request.Request(
        UNITRUST_URL,
        headers={"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"}
    )
    with opener.open(req, timeout=20) as r:
        raw = r.read()
    page = raw.decode("utf-8", errors="replace")

    idx = page.find("DetailCode")
    if idx < 0:
        raise ValueError("找不到 DetailCode，頁面格式可能已變更")

    start = max(0, idx - 5000)
    chunk = page[start:idx + 200]
    arr_start = chunk.rfind("[{")
    if arr_start < 0:
        raise ValueError("找不到 JSON 陣列起點 '[{'")

    raw_chunk = page[start + arr_start:]
    depth, end = 0, 0
    for i, ch in enumerate(raw_chunk):
        if ch == "[":
            depth += 1
        elif ch == "]":
            depth -= 1
            if depth == 0:
                end = i + 1
                break
    if not end:
        raise ValueError("JSON 陣列括號不匹配")

    json_str = raw_chunk[:end].replace("&quot;", '"').replace("&#34;", '"').replace("&amp;", "&")
    data = json.loads(json_str)

    stocks = []
    date_str = ""
    for d in data:
        code = str(d.get("DetailCode", "")).strip()
        if not code or len(code) != 4:
            continue
        if not date_str and d.get("TranDate"):
            date_str = d["TranDate"][:10].replace("-", "/")
        stocks.append({
            "code":   code,
            "name":   str(d.get("DetailName", "")).strip(),
            "shares": int(d.get("Share", 0)),
            "weight": round(float(d.get("NavRate", 0)), 4),
        })
    return date_str, stocks


def download_today():
    today = taipei_today()
    if not is_weekday(today):
        print(f"今天是週末（{today}），跳過下載。")
        return
    date_str = today.strftime("%Y%m%d")
    date_fmt  = today.strftime("%Y_%m_%d")
    print(f"📥  下載 {today} 資料…")

    # ── 復華 ──
    fuhua_dest = os.path.join(FUHUA_DIR, f"復華台灣未來50主動式ETF基金-基金資產-{date_fmt}.xlsx")
    if os.path.exists(fuhua_dest):
        print(f"  (復華 {today} 已存在，跳過)")
    else:
        url = FUHUA_API.format(date=date_str)
        download_file(url, fuhua_dest)

    # ── 群益 00992A（JSON API）──
    qunyi_dest = os.path.join(QUNYI_DIR, f"{today.strftime('%Y%m%d')}.xlsx")
    if os.path.exists(qunyi_dest):
        print(f"  (群益992 {today} 已存在，跳過)")
    else:
        try:
            date_api = today.strftime("%Y/%m/%d")
            print(f"  📡  群益992 API 取得 {date_api}…")
            stocks = fetch_qunyi_api(date_api, fund_id="500")
            print(f"  ✓ 取得 {len(stocks)} 檔")
            _save_qunyi_xlsx(stocks, today, qunyi_dest)
        except Exception as e:
            print(f"  ⚠️  群益992 資料取得失敗：{e}")

    # ── 群益 00982A（JSON API）──
    qunyi982_dest = os.path.join(QUNYI982_DIR, f"{today.strftime('%Y%m%d')}.xlsx")
    if os.path.exists(qunyi982_dest):
        print(f"  (群益982 {today} 已存在，跳過)")
    else:
        try:
            date_api = today.strftime("%Y/%m/%d")
            print(f"  📡  群益982 API 取得 {date_api}…")
            stocks = fetch_qunyi_api(date_api, fund_id="399")
            print(f"  ✓ 取得 {len(stocks)} 檔")
            _save_qunyi_xlsx(stocks, today, qunyi982_dest)
        except Exception as e:
            print(f"  ⚠️  群益982 資料取得失敗：{e}")

    # ── 統一（ezmoney 頁面）──
    unitrust_dest = os.path.join(UNITRUST_DIR, f"{today.strftime('%Y%m%d')}.json")
    today_str = today.strftime("%Y/%m/%d")
    # 若檔案存在但日期不對（早上抓到舊資料），刪掉重抓
    if os.path.exists(unitrust_dest):
        cached = json.load(open(unitrust_dest, encoding="utf-8"))
        if cached.get("date") != today_str:
            os.remove(unitrust_dest)
            print(f"  (統一 快取日期={cached.get('date')}，重新抓取)")
        else:
            print(f"  (統一 {today} 已存在，跳過)")
    if not os.path.exists(unitrust_dest):
        try:
            print(f"  📡  統一 ezmoney 取得…")
            date_api, stocks = fetch_unitrust_ezmoney()
            if stocks:
                with open(unitrust_dest, "w", encoding="utf-8") as f:
                    json.dump({"date": date_api, "stocks": stocks}, f, ensure_ascii=False)
                print(f"  ✓ 取得 {len(stocks)} 檔（{date_api}）→ {unitrust_dest}")
            else:
                print(f"  ⚠️  統一資料為空")
        except Exception as e:
            print(f"  ⚠️  統一資料取得失敗：{e}")

    # ── 野村 00980A（JSON API）──
    nomura_dest = os.path.join(NOMURA_DIR, f"{today.strftime('%Y%m%d')}.json")
    # 若檔案存在但日期不對（早上抓到舊資料），刪掉重抓
    if os.path.exists(nomura_dest):
        cached = json.load(open(nomura_dest, encoding="utf-8"))
        if cached.get("date") != today_str:
            os.remove(nomura_dest)
            print(f"  (野村 快取日期={cached.get('date')}，重新抓取)")
        else:
            print(f"  (野村 {today} 已存在，跳過)")
    if not os.path.exists(nomura_dest):
        try:
            print(f"  📡  野村 API 取得…")
            date_api, stocks = fetch_nomura_api()
            if stocks:
                with open(nomura_dest, "w", encoding="utf-8") as f:
                    json.dump({"date": date_api, "stocks": stocks}, f, ensure_ascii=False)
                print(f"  ✓ 取得 {len(stocks)} 檔（{date_api}）→ {nomura_dest}")
        except Exception as e:
            print(f"  ⚠️  野村資料取得失敗：{e}")


def _save_qunyi_xlsx(stocks: list, date: datetime.date, dest: str):
    """將群益持股清單存成 xlsx（與手動下載格式相同）。"""
    wb = openpyxl.Workbook()
    ws_info = wb.active
    ws_info.title = "投資組合"
    ws_info.append(["基金淨資產價值(元)", ""])
    ws_info.append(["每受益權單位淨資產價值(元)-台幣交易", ""])
    ws_info.append(["已發行受益權單位總數-台幣交易", ""])

    ws = wb.create_sheet("股票")
    ws.append(["股票代號", "股票名稱", "持股權重(%)", "股數"])
    for s in stocks:
        ws.append([s["code"], s["name"], f"{s['weight']}%", s["shares"]])

    wb.create_sheet("其他資產").append(["現金", ""])
    wb.save(dest)
    print(f"  ✓ 群益資料已存至 {dest}")


# ── 解析函式 ─────────────────────────────────────────────────────────────
def parse_fuhua(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    rows = [r for r in ws.iter_rows(values_only=True) if any(v is not None for v in r)]
    date_str = ""
    for row in rows:
        if row[0] and isinstance(row[0], str) and "日期" in row[0]:
            date_str = row[0].replace("日期: ", "").strip()
            break
    stocks = []
    in_data = False
    for row in rows:
        if row[0] == "證券代號":
            in_data = True
            continue
        if in_data and row[0]:
            shares_str = str(row[2] or "0").replace(",", "")
            weight_str = str(row[4] or "0%").replace("%", "")
            stocks.append({
                "code":   str(row[0]).strip(),
                "name":   str(row[1] or "").strip(),
                "shares": int(shares_str),
                "weight": float(weight_str),
            })
    return date_str, stocks


def parse_qunyi(filepath):
    basename = os.path.basename(filepath).replace(".xlsx", "")
    try:
        if len(basename) == 8 and basename.isdigit():
            d = datetime.datetime.strptime(basename, "%Y%m%d")
            date_str = d.strftime("%Y/%m/%d")
        else:
            day = int(basename)
            mtime = os.path.getmtime(filepath)
            file_date = datetime.datetime.fromtimestamp(mtime, TAIPEI_TZ)
            date_str = f"{file_date.year}/{file_date.month:02d}/{day:02d}"
    except ValueError:
        date_str = basename

    wb = openpyxl.load_workbook(filepath)
    ws = wb["股票"]
    stocks = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        shares_str = str(row[3] or "0").replace(",", "")
        weight_str = str(row[2] or "0%").replace("%", "")
        stocks.append({
            "code":   str(row[0]).strip(),
            "name":   str(row[1] or "").strip(),
            "shares": int(float(shares_str)),
            "weight": float(weight_str),
        })
    return date_str, stocks


def parse_unitrust(filepath):
    with open(filepath, encoding="utf-8") as f:
        obj = json.load(f)
    return obj["date"], obj["stocks"]


def load_all_data():
    fuhua_data, qunyi_data, qunyi982_data, unitrust_data, nomura_data = {}, {}, {}, {}, {}

    for f in sorted(glob.glob(os.path.join(FUHUA_DIR, "*.xlsx"))):
        try:
            d, s = parse_fuhua(f)
            if d: fuhua_data[d] = s
        except Exception as e:
            print(f"  ⚠️  解析復華失敗 {f}: {e}")

    for f in sorted(glob.glob(os.path.join(QUNYI_DIR, "*.xlsx"))):
        try:
            d, s = parse_qunyi(f)
            if d: qunyi_data[d] = s
        except Exception as e:
            print(f"  ⚠️  解析群益992失敗 {f}: {e}")

    for f in sorted(glob.glob(os.path.join(QUNYI982_DIR, "*.xlsx"))):
        try:
            d, s = parse_qunyi(f)
            if d: qunyi982_data[d] = s
        except Exception as e:
            print(f"  ⚠️  解析群益982失敗 {f}: {e}")

    for f in sorted(glob.glob(os.path.join(UNITRUST_DIR, "*.json"))):
        try:
            d, s = parse_unitrust(f)
            if d: unitrust_data[d] = s
        except Exception as e:
            print(f"  ⚠️  解析統一失敗 {f}: {e}")

    for f in sorted(glob.glob(os.path.join(NOMURA_DIR, "*.json"))):
        try:
            d, s = parse_unitrust(f)
            if d: nomura_data[d] = s
        except Exception as e:
            print(f"  ⚠️  解析野村失敗 {f}: {e}")

    return fuhua_data, qunyi_data, qunyi982_data, unitrust_data, nomura_data


# ── 產生 HTML ─────────────────────────────────────────────────────────────
def generate_html(fuhua_data, qunyi_data, qunyi982_data, unitrust_data, nomura_data) -> str:
    now_taipei = datetime.datetime.now(TAIPEI_TZ)
    update_time = now_taipei.strftime("%Y/%m/%d %H:%M 更新")

    fuhua_js     = json.dumps(fuhua_data, ensure_ascii=False)
    qunyi_js     = json.dumps(qunyi_data, ensure_ascii=False)
    qunyi982_js  = json.dumps(qunyi982_data, ensure_ascii=False)
    unitrust_js  = json.dumps(unitrust_data, ensure_ascii=False)
    nomura_js    = json.dumps(nomura_data, ensure_ascii=False)

    return f'''<!DOCTYPE html>
<html lang="zh-TW">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0">
<title>主動式ETF 持股追蹤</title>
<style>
  *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}
  :root {{
    --bg: #f5f5f7; --card: #ffffff; --text: #1d1d1f; --sub: #6e6e73;
    --border: #e0e0e5; --accent: #0071e3; --badge-bg: #f0f0f5;
    --add: #16a34a; --add-bg: #dcfce7;
    --remove: #dc2626; --remove-bg: #fee2e2;
    --up: #ea580c; --up-bg: #ffedd5;
    --down: #2563eb; --down-bg: #dbeafe;
  }}
  body {{ background: var(--bg); color: var(--text); font-family: -apple-system, "PingFang TC", sans-serif; font-size: 15px; line-height: 1.5; }}

  header {{ background: var(--card); border-bottom: 1px solid var(--border); padding: 14px 16px 10px; position: sticky; top: 0; z-index: 100; }}
  header h1 {{ font-size: 17px; font-weight: 700; }}
  header p {{ font-size: 12px; color: var(--sub); margin-top: 2px; }}

  .tabs {{ display: flex; background: var(--card); border-bottom: 1px solid var(--border); overflow-x: auto; scrollbar-width: none; position: sticky; top: 49px; z-index: 99; }}
  .tabs::-webkit-scrollbar {{ display: none; }}
  .tab {{ flex: 1; min-width: 90px; padding: 10px 12px; text-align: center; font-size: 13px; font-weight: 600; cursor: pointer; border-bottom: 2px solid transparent; color: var(--sub); transition: all 0.2s; white-space: nowrap; }}
  .tab.active {{ color: var(--accent); border-bottom-color: var(--accent); }}

  .panel {{ display: none; padding: 12px; max-width: 600px; margin: 0 auto; }}
  .panel.active {{ display: block; }}

  .date-selector {{ background: var(--card); border-radius: 12px; padding: 12px 14px; margin-bottom: 12px; border: 1px solid var(--border); }}
  .ds-label {{ font-size: 11px; font-weight: 700; color: var(--sub); text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 8px; }}
  .ds-row {{ display: flex; align-items: center; gap: 8px; flex-wrap: wrap; }}
  .ds-select {{ flex: 1; min-width: 110px; padding: 7px 28px 7px 10px; border: 1px solid var(--border); border-radius: 8px; font-size: 13px; font-family: inherit; background: var(--bg); color: var(--text); appearance: none; background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='12' height='12' viewBox='0 0 24 24' fill='none' stroke='%236e6e73' stroke-width='2'%3E%3Cpath d='M6 9l6 6 6-6'/%3E%3C/svg%3E"); background-repeat: no-repeat; background-position: right 8px center; }}
  .ds-arrow {{ color: var(--sub); font-size: 14px; flex-shrink: 0; }}
  .ds-period {{ font-size: 12px; color: var(--sub); margin-top: 6px; }}

  .section {{ background: var(--card); border-radius: 12px; border: 1px solid var(--border); margin-bottom: 12px; overflow: hidden; }}
  .section-title {{ padding: 11px 14px 9px; font-size: 12px; font-weight: 700; color: var(--sub); text-transform: uppercase; letter-spacing: 0.5px; border-bottom: 1px solid var(--border); display: flex; justify-content: space-between; align-items: center; }}
  .section-count {{ font-size: 11px; background: var(--badge-bg); color: var(--sub); padding: 2px 7px; border-radius: 10px; font-weight: 600; }}

  .summary-row {{ display: flex; gap: 6px; padding: 10px 14px; flex-wrap: wrap; }}
  .s-badge {{ display: flex; align-items: center; gap: 4px; padding: 5px 10px; border-radius: 8px; font-size: 12px; font-weight: 700; }}
  .s-badge .label {{ font-weight: 400; font-size: 11px; }}
  .s-add {{ background: var(--add-bg); color: var(--add); }}
  .s-remove {{ background: var(--remove-bg); color: var(--remove); }}
  .s-up {{ background: var(--up-bg); color: var(--up); }}
  .s-down {{ background: var(--down-bg); color: var(--down); }}

  .change-list {{ padding: 4px 0; }}
  .change-card {{ padding: 10px 14px; border-bottom: 1px solid var(--border); display: grid; grid-template-columns: auto 1fr auto; gap: 0 10px; align-items: start; }}
  .change-card:last-child {{ border-bottom: none; }}

  .op-badge {{ display: inline-flex; align-items: center; padding: 2px 8px; border-radius: 6px; font-size: 11px; font-weight: 700; white-space: nowrap; margin-top: 2px; }}
  .op-add {{ background: var(--add-bg); color: var(--add); }}
  .op-remove {{ background: var(--remove-bg); color: var(--remove); }}
  .op-up {{ background: var(--up-bg); color: var(--up); }}
  .op-down {{ background: var(--down-bg); color: var(--down); }}

  .cc-main {{ display: flex; flex-direction: column; gap: 1px; }}
  .cc-title {{ font-size: 14px; font-weight: 600; }}
  .cc-code {{ font-size: 11px; color: var(--sub); }}
  .cc-details {{ display: flex; flex-wrap: wrap; gap: 6px; margin-top: 4px; }}
  .cc-detail {{ font-size: 11px; color: var(--sub); }}
  .cc-detail strong {{ color: var(--text); font-weight: 600; }}
  .cc-detail.hi-up {{ color: var(--up); }} .cc-detail.hi-up strong {{ color: var(--up); }}
  .cc-detail.hi-down {{ color: var(--down); }} .cc-detail.hi-down strong {{ color: var(--down); }}
  .cc-detail.hi-add {{ color: var(--add); }} .cc-detail.hi-add strong {{ color: var(--add); }}
  .cc-detail.hi-remove {{ color: var(--remove); }} .cc-detail.hi-remove strong {{ color: var(--remove); }}

  .cc-weight {{ text-align: right; }}
  .cc-w-val {{ font-size: 14px; font-weight: 700; }}
  .cc-w-diff {{ font-size: 11px; margin-top: 1px; font-weight: 600; }}
  .w-up {{ color: var(--up); }} .w-down {{ color: var(--down); }}
  .w-add {{ color: var(--add); }} .w-remove {{ color: var(--remove); }}

  .holdings-table {{ width: 100%; }}
  .holdings-row {{ display: flex; align-items: center; padding: 7px 14px; border-bottom: 1px solid var(--border); gap: 8px; }}
  .holdings-row:last-child {{ border-bottom: none; }}
  .h-rank {{ font-size: 11px; color: var(--sub); min-width: 20px; }}
  .h-code {{ font-size: 11px; color: var(--sub); min-width: 38px; }}
  .h-name {{ font-size: 13px; flex: 1; font-weight: 500; }}
  .h-shares {{ font-size: 11px; color: var(--sub); text-align: right; min-width: 70px; }}
  .h-bar-wrap {{ width: 40px; flex-shrink: 0; }}
  .h-bar {{ height: 3px; background: #e0e0e5; border-radius: 2px; overflow: hidden; }}
  .h-bar-fill {{ height: 100%; background: var(--accent); border-radius: 2px; }}
  .h-weight {{ font-size: 12px; font-weight: 700; min-width: 44px; text-align: right; color: var(--accent); }}

  .toggle-btn {{ background: none; border: 1px solid var(--border); border-radius: 6px; padding: 3px 8px; font-size: 11px; cursor: pointer; color: var(--accent); font-family: inherit; }}
  .empty-msg {{ padding: 14px; text-align: center; font-size: 13px; color: var(--sub); }}
  .fund-meta {{ background: var(--card); border-radius: 12px; padding: 12px 14px; margin-bottom: 12px; border: 1px solid var(--border); }}
  .fund-name {{ font-size: 16px; font-weight: 700; }}
  .fund-sub {{ font-size: 12px; color: var(--sub); margin-top: 3px; }}

  .search-wrap {{ position: relative; margin-bottom: 12px; }}
  .search-input {{ width: 100%; padding: 10px 40px 10px 14px; border: 1px solid var(--border); border-radius: 12px; font-size: 14px; font-family: inherit; background: var(--card); color: var(--text); outline: none; -webkit-appearance: none; }}
  .search-input:focus {{ border-color: var(--accent); box-shadow: 0 0 0 3px rgba(0,113,227,0.12); }}
  .search-icon {{ position: absolute; right: 13px; top: 50%; transform: translateY(-50%); color: var(--sub); pointer-events: none; font-size: 16px; }}
  .search-dropdown {{ position: absolute; top: calc(100% + 4px); left: 0; right: 0; background: var(--card); border: 1px solid var(--border); border-radius: 12px; overflow: hidden; z-index: 50; box-shadow: 0 4px 24px rgba(0,0,0,0.10); }}
  .sr-item {{ padding: 10px 14px; cursor: pointer; display: flex; justify-content: space-between; align-items: center; font-size: 14px; border-bottom: 1px solid var(--border); }}
  .sr-item:last-child {{ border-bottom: none; }}
  .sr-item:active {{ background: var(--bg); }}
  .sr-code {{ font-size: 12px; color: var(--sub); }}
  .sr-empty {{ padding: 12px 14px; font-size: 13px; color: var(--sub); text-align: center; }}

  .stock-history {{ background: var(--card); border-radius: 12px; border: 1px solid var(--border); margin-bottom: 12px; overflow: hidden; }}
  .sh-header {{ padding: 11px 14px; display: flex; justify-content: space-between; align-items: flex-start; border-bottom: 1px solid var(--border); }}
  .sh-title {{ font-size: 14px; font-weight: 700; }}
  .sh-subtitle {{ font-size: 11px; color: var(--sub); margin-top: 2px; }}
  .sh-close {{ background: none; border: none; font-size: 20px; cursor: pointer; color: var(--sub); padding: 0; line-height: 1; margin-left: 8px; flex-shrink: 0; }}
  .sh-row {{ display: grid; grid-template-columns: 40px 42px 1fr auto auto; gap: 6px; align-items: center; padding: 9px 14px; border-bottom: 1px solid var(--border); }}
  .sh-row:last-child {{ border-bottom: none; }}
  .sh-date {{ font-size: 12px; color: var(--sub); font-weight: 500; }}
  .sh-shares {{ font-size: 13px; font-weight: 500; }}
  .sh-weight {{ font-size: 13px; font-weight: 700; color: var(--accent); text-align: right; }}
  .sh-diff {{ font-size: 11px; font-weight: 600; text-align: right; white-space: nowrap; }}
  .sh-none {{ font-size: 12px; color: var(--sub); font-style: italic; grid-column: 3/-1; }}
</style>
</head>
<body>

<header>
  <h1>主動式ETF 持股追蹤</h1>
  <p>{update_time}</p>
</header>

<div class="tabs">
  <div class="tab active" onclick="switchTab(\'t991a\')">00991A<br><small style="font-weight:400;font-size:10px">復華未來50</small></div>
  <div class="tab" onclick="switchTab(\'t992a\')">00992A<br><small style="font-weight:400;font-size:10px">群益科技創新</small></div>
  <div class="tab" onclick="switchTab(\'t982a\')">00982A<br><small style="font-weight:400;font-size:10px">群益科技高息</small></div>
  <div class="tab" onclick="switchTab(\'t981a\')">00981A<br><small style="font-weight:400;font-size:10px">統一台股增長</small></div>
  <div class="tab" onclick="switchTab(\'t980a\')">00980A<br><small style="font-weight:400;font-size:10px">野村台灣創新</small></div>
</div>

<!-- 00991A -->
<div id="t991a" class="panel active">
  <div class="fund-meta">
    <div class="fund-name">00991A 復華台灣未來50</div>
    <div class="fund-sub">主動式ETF ｜ 復華投信</div>
  </div>
  <div class="search-wrap">
    <input class="search-input" id="srch991a" placeholder="搜尋成分股名稱或代號…" autocomplete="off" oninput="searchStock(\'991a\',this.value)" onblur="hideDropdown(\'991a\')" onfocus="searchStock(\'991a\',this.value)">
    <span class="search-icon">⌕</span>
    <div class="search-dropdown" id="drop991a" style="display:none"></div>
  </div>
  <div class="stock-history" id="hist991a" style="display:none"></div>
  <div class="date-selector">
    <div class="ds-label">選擇比較日期</div>
    <div class="ds-row">
      <select class="ds-select" id="sel991a-from" onchange="render(\'991a\')"></select>
      <span class="ds-arrow">→</span>
      <select class="ds-select" id="sel991a-to" onchange="render(\'991a\')"></select>
    </div>
    <div class="ds-period" id="period991a"></div>
  </div>
  <div class="section">
    <div class="section-title">變動摘要 <span class="section-count" id="cnt991a">—</span></div>
    <div class="summary-row" id="sum991a"></div>
  </div>
  <div class="section">
    <div class="section-title">持股變動明細</div>
    <div class="change-list" id="chg991a"></div>
  </div>
  <div class="section">
    <div class="section-title">完整持股 <button class="toggle-btn" onclick="toggleEl(\'full991a\',this)">展開</button></div>
    <div id="full991a" class="holdings-table" style="display:none"></div>
  </div>
</div>

<!-- 00992A -->
<div id="t992a" class="panel">
  <div class="fund-meta">
    <div class="fund-name">00992A 群益科技創新</div>
    <div class="fund-sub">主動式ETF ｜ 群益投信</div>
  </div>
  <div class="search-wrap">
    <input class="search-input" id="srch992a" placeholder="搜尋成分股名稱或代號…" autocomplete="off" oninput="searchStock(\'992a\',this.value)" onblur="hideDropdown(\'992a\')" onfocus="searchStock(\'992a\',this.value)">
    <span class="search-icon">⌕</span>
    <div class="search-dropdown" id="drop992a" style="display:none"></div>
  </div>
  <div class="stock-history" id="hist992a" style="display:none"></div>
  <div class="date-selector">
    <div class="ds-label">選擇比較日期</div>
    <div class="ds-row">
      <select class="ds-select" id="sel992a-from" onchange="render(\'992a\')"></select>
      <span class="ds-arrow">→</span>
      <select class="ds-select" id="sel992a-to" onchange="render(\'992a\')"></select>
    </div>
    <div class="ds-period" id="period992a"></div>
  </div>
  <div class="section">
    <div class="section-title">變動摘要 <span class="section-count" id="cnt992a">—</span></div>
    <div class="summary-row" id="sum992a"></div>
  </div>
  <div class="section">
    <div class="section-title">持股變動明細</div>
    <div class="change-list" id="chg992a"></div>
  </div>
  <div class="section">
    <div class="section-title">完整持股 <button class="toggle-btn" onclick="toggleEl(\'full992a\',this)">展開</button></div>
    <div id="full992a" class="holdings-table" style="display:none"></div>
  </div>
</div>

<!-- 00982A -->
<div id="t982a" class="panel">
  <div class="fund-meta">
    <div class="fund-name">00982A 群益科技高息成長</div>
    <div class="fund-sub">主動式ETF ｜ 群益投信</div>
  </div>
  <div class="search-wrap">
    <input class="search-input" id="srch982a" placeholder="搜尋成分股名稱或代號…" autocomplete="off" oninput="searchStock(\'982a\',this.value)" onblur="hideDropdown(\'982a\')" onfocus="searchStock(\'982a\',this.value)">
    <span class="search-icon">⌕</span>
    <div class="search-dropdown" id="drop982a" style="display:none"></div>
  </div>
  <div class="stock-history" id="hist982a" style="display:none"></div>
  <div class="date-selector">
    <div class="ds-label">選擇比較日期</div>
    <div class="ds-row">
      <select class="ds-select" id="sel982a-from" onchange="render(\'982a\')"></select>
      <span class="ds-arrow">→</span>
      <select class="ds-select" id="sel982a-to" onchange="render(\'982a\')"></select>
    </div>
    <div class="ds-period" id="period982a"></div>
  </div>
  <div class="section">
    <div class="section-title">變動摘要 <span class="section-count" id="cnt982a">—</span></div>
    <div class="summary-row" id="sum982a"></div>
  </div>
  <div class="section">
    <div class="section-title">持股變動明細</div>
    <div class="change-list" id="chg982a"></div>
  </div>
  <div class="section">
    <div class="section-title">完整持股 <button class="toggle-btn" onclick="toggleEl(\'full982a\',this)">展開</button></div>
    <div id="full982a" class="holdings-table" style="display:none"></div>
  </div>
</div>

<!-- 00981A -->
<div id="t981a" class="panel">
  <div class="fund-meta">
    <div class="fund-name">00981A 統一台股增長</div>
    <div class="fund-sub">主動式ETF ｜ 統一投信</div>
  </div>
  <div class="search-wrap">
    <input class="search-input" id="srch981a" placeholder="搜尋成分股名稱或代號…" autocomplete="off" oninput="searchStock(\'981a\',this.value)" onblur="hideDropdown(\'981a\')" onfocus="searchStock(\'981a\',this.value)">
    <span class="search-icon">⌕</span>
    <div class="search-dropdown" id="drop981a" style="display:none"></div>
  </div>
  <div class="stock-history" id="hist981a" style="display:none"></div>
  <div class="date-selector">
    <div class="ds-label">選擇比較日期</div>
    <div class="ds-row">
      <select class="ds-select" id="sel981a-from" onchange="render(\'981a\')"></select>
      <span class="ds-arrow">→</span>
      <select class="ds-select" id="sel981a-to" onchange="render(\'981a\')"></select>
    </div>
    <div class="ds-period" id="period981a"></div>
  </div>
  <div class="section">
    <div class="section-title">變動摘要 <span class="section-count" id="cnt981a">—</span></div>
    <div class="summary-row" id="sum981a"></div>
  </div>
  <div class="section">
    <div class="section-title">持股變動明細</div>
    <div class="change-list" id="chg981a"></div>
  </div>
  <div class="section">
    <div class="section-title">完整持股 <button class="toggle-btn" onclick="toggleEl(\'full981a\',this)">展開</button></div>
    <div id="full981a" class="holdings-table" style="display:none"></div>
  </div>
</div>

<!-- 00980A -->
<div id="t980a" class="panel">
  <div class="fund-meta">
    <div class="fund-name">00980A 野村台灣創新科技50</div>
    <div class="fund-sub">主動式ETF ｜ 野村投信</div>
  </div>
  <div class="search-wrap">
    <input class="search-input" id="srch980a" placeholder="搜尋成分股名稱或代號…" autocomplete="off" oninput="searchStock(\'980a\',this.value)" onblur="hideDropdown(\'980a\')" onfocus="searchStock(\'980a\',this.value)">
    <span class="search-icon">⌕</span>
    <div class="search-dropdown" id="drop980a" style="display:none"></div>
  </div>
  <div class="stock-history" id="hist980a" style="display:none"></div>
  <div class="date-selector">
    <div class="ds-label">選擇比較日期</div>
    <div class="ds-row">
      <select class="ds-select" id="sel980a-from" onchange="render(\'980a\')"></select>
      <span class="ds-arrow">→</span>
      <select class="ds-select" id="sel980a-to" onchange="render(\'980a\')"></select>
    </div>
    <div class="ds-period" id="period980a"></div>
  </div>
  <div class="section">
    <div class="section-title">變動摘要 <span class="section-count" id="cnt980a">—</span></div>
    <div class="summary-row" id="sum980a"></div>
  </div>
  <div class="section">
    <div class="section-title">持股變動明細</div>
    <div class="change-list" id="chg980a"></div>
  </div>
  <div class="section">
    <div class="section-title">完整持股 <button class="toggle-btn" onclick="toggleEl(\'full980a\',this)">展開</button></div>
    <div id="full980a" class="holdings-table" style="display:none"></div>
  </div>
</div>

<script>
const DB = {{
  "991a": {fuhua_js},
  "992a": {qunyi_js},
  "982a": {qunyi982_js},
  "981a": {unitrust_js},
  "980a": {nomura_js}
}};

function fmtShares(n) {{ return n.toLocaleString("zh-TW"); }}
function fmtDate(d) {{ return d.replace("2026/","").replace(/\\/0/g,"/"); }}
function sortedDates(etf) {{ return Object.keys(DB[etf]).sort(); }}

function computeChanges(from, to) {{
  const fMap={{}}, tMap={{}};
  from.forEach(s=>fMap[s.code]=s); to.forEach(s=>tMap[s.code]=s);
  const codes=new Set([...Object.keys(fMap),...Object.keys(tMap)]);
  const changes=[];
  codes.forEach(c=>{{
    const f=fMap[c], t=tMap[c];
    const fS=f?f.shares:0, tS=t?t.shares:0;
    if(fS===tS) return;
    const fW=f?f.weight:0, tW=t?t.weight:0;
    const op=fS===0?"add":tS===0?"remove":fS<tS?"up":"down";
    changes.push({{code:c,name:(t||f).name,op,diff:tS-fS,fShares:fS,tShares:tS,fWeight:fW,tWeight:tW,wDiff:+(tW-fW).toFixed(3)}});
  }});
  const ord={{add:0,remove:1,up:2,down:3}};
  changes.sort((a,b)=>{{
    const wa=Math.abs(a.wDiff),wb=Math.abs(b.wDiff);
    if(Math.abs(wa-wb)>0.05) return wb-wa;
    return ord[a.op]-ord[b.op];
  }});
  return changes;
}}

const OP_LABEL={{add:"建倉",remove:"清倉",up:"加碼",down:"減碼"}};
const OP_CLS={{add:"op-add",remove:"op-remove",up:"op-up",down:"op-down"}};
const W_CLS={{add:"w-add",remove:"w-remove",up:"w-up",down:"w-down"}};
const HI_CLS={{add:"hi-add",remove:"hi-remove",up:"hi-up",down:"hi-down"}};

function render(etfId) {{
  const from = document.getElementById("sel"+etfId+"-from").value;
  const to   = document.getElementById("sel"+etfId+"-to").value;
  const fromS = DB[etfId][from]||[], toS = DB[etfId][to]||[];
  const changes = computeChanges(fromS, toS);

  const cnt={{add:0,remove:0,up:0,down:0}};
  changes.forEach(c=>cnt[c.op]++);

  document.getElementById("period"+etfId).textContent = fmtDate(from)+" → "+fmtDate(to);
  document.getElementById("cnt"+etfId).textContent = changes.length+" 檔";
  document.getElementById("sum"+etfId).innerHTML =
    `<span class="s-badge s-add"><span>${{cnt.add}}</span><span class="label">建倉</span></span>
     <span class="s-badge s-remove"><span>${{cnt.remove}}</span><span class="label">清倉</span></span>
     <span class="s-badge s-up"><span>${{cnt.up}}</span><span class="label">加碼</span></span>
     <span class="s-badge s-down"><span>${{cnt.down}}</span><span class="label">減碼</span></span>`;

  const chgEl = document.getElementById("chg"+etfId);
  if(!changes.length){{chgEl.innerHTML=\'<div class="empty-msg">此期間無持股變動</div>\';}}
  else {{
    chgEl.innerHTML = changes.map(c=>{{
      const sign=c.diff>0?"+":"";
      const wSign=c.wDiff>0?"+":"";
      let details="";
      if(c.op==="add")   details+=`<span class="cc-detail hi-add"><strong>+${{fmtShares(c.tShares)}} 股</strong></span>`;
      if(c.op==="remove") details+=`<span class="cc-detail hi-remove"><strong>-${{fmtShares(c.fShares)}} 股</strong></span>`;
      if(c.op==="up"||c.op==="down") details+=`<span class="cc-detail ${{HI_CLS[c.op]}}"><strong>${{sign}}${{fmtShares(c.diff)}} 股</strong></span>`;
      if(c.op!=="remove") details+=`<span class="cc-detail">持股 <strong>${{fmtShares(c.tShares)}}</strong> 股</span>`;
      if(c.op==="up"||c.op==="down") details+=`<span class="cc-detail">權重 <strong>${{c.fWeight.toFixed(2)}}% → ${{c.tWeight.toFixed(2)}}%</strong></span>`;
      if(c.op==="add") details+=`<span class="cc-detail">權重 <strong>${{c.tWeight.toFixed(2)}}%</strong></span>`;
      const wDisp = c.op==="remove"
        ? `<div class="cc-w-val w-remove">${{c.fWeight.toFixed(2)}}%</div><div class="cc-w-diff w-remove">→ 清倉</div>`
        : `<div class="cc-w-val">${{c.tWeight.toFixed(2)}}%</div>${{(c.op==="up"||c.op==="down")?`<div class="cc-w-diff ${{W_CLS[c.op]}}">${{wSign}}${{c.wDiff.toFixed(2)}}%</div>`:"" }}`;
      return `<div class="change-card">
        <div><span class="op-badge ${{OP_CLS[c.op]}}">${{OP_LABEL[c.op]}}</span></div>
        <div class="cc-main">
          <div class="cc-title">${{c.name}}</div><div class="cc-code">${{c.code}}</div>
          <div class="cc-details">${{details}}</div>
        </div>
        <div class="cc-weight">${{wDisp}}</div>
      </div>`;
    }}).join("");
  }}
  renderFull(etfId, toS);
}}

function renderFull(etfId, stocks) {{
  const el=document.getElementById("full"+etfId); if(!el) return;
  const maxW=Math.max(...stocks.map(s=>s.weight));
  el.innerHTML=stocks.map((s,i)=>`
    <div class="holdings-row">
      <span class="h-rank">${{i+1}}</span><span class="h-code">${{s.code}}</span>
      <span class="h-name">${{s.name}}</span><span class="h-shares">${{fmtShares(s.shares)}}</span>
      <div class="h-bar-wrap"><div class="h-bar"><div class="h-bar-fill" style="width:${{maxW>0?(s.weight/maxW*100).toFixed(1):0}}%"></div></div></div>
      <span class="h-weight">${{s.weight.toFixed(2)}}%</span>
    </div>`).join("");
}}

function initSelects(etfId) {{
  const dates=sortedDates(etfId);
  const fEl=document.getElementById("sel"+etfId+"-from"), tEl=document.getElementById("sel"+etfId+"-to");
  dates.forEach(d=>{{ fEl.appendChild(new Option(fmtDate(d),d)); tEl.appendChild(new Option(fmtDate(d),d)); }});
  if(dates.length>=2){{ fEl.value=dates[dates.length-2]; tEl.value=dates[dates.length-1]; }}
  else if(dates.length===1){{ fEl.value=tEl.value=dates[0]; }}
}}

function toggleEl(id,btn) {{
  const el=document.getElementById(id), h=el.style.display==="none";
  el.style.display=h?"block":"none"; btn.textContent=h?"收合":"展開";
}}
function switchTab(tid) {{
  document.querySelectorAll(".tab").forEach(t=>t.classList.toggle("active",t.getAttribute("onclick").includes(tid)));
  document.querySelectorAll(".panel").forEach(p=>p.classList.toggle("active",p.id===tid));
}}

// ── 搜尋功能 ────────────────────────────────────────────────────────────
function searchStock(etfId, query) {{
  const drop = document.getElementById("drop"+etfId);
  query = query.trim();
  if(!query) {{ drop.style.display="none"; return; }}

  const allStocks = {{}};
  Object.values(DB[etfId]).forEach(list => list.forEach(s => {{ allStocks[s.code]=s.name; }}));
  const q = query.toLowerCase();
  const matches = Object.entries(allStocks)
    .filter(([code,name]) => code.startsWith(q) || name.includes(query))
    .sort((a,b) => (a[0].startsWith(q)?-1:0)-(b[0].startsWith(q)?-1:0))
    .slice(0,8);

  if(!matches.length) {{
    drop.innerHTML = '<div class="sr-empty">找不到相符股票</div>';
  }} else {{
    drop.innerHTML = matches.map(([code,name]) =>
      `<div class="sr-item" onmousedown="showStockHistory('${{etfId}}','${{code}}','${{name}}')">`+
      `<span>${{name}}</span><span class="sr-code">${{code}}</span></div>`
    ).join("");
  }}
  drop.style.display = "block";
}}

function hideDropdown(etfId) {{
  setTimeout(() => {{ document.getElementById("drop"+etfId).style.display="none"; }}, 150);
}}

function showStockHistory(etfId, code, name) {{
  document.getElementById("drop"+etfId).style.display = "none";
  document.getElementById("srch"+etfId).value = name+" "+code;

  const HIST_DAYS = 5;
  const dates = sortedDates(etfId).slice(-HIST_DAYS);
  const histEl = document.getElementById("hist"+etfId);

  const rows = dates.map((d, i) => {{
    const s    = (DB[etfId][d]||[]).find(x=>x.code===code);
    const prev = i>0 ? (DB[etfId][dates[i-1]]||[]).find(x=>x.code===code) : null;

    const shares = s ? s.shares : null;
    const weight = s ? s.weight : null;
    const prevShares = prev ? prev.shares : null;

    let op=null, diff=null;
    if(shares!==null && prevShares===null)      op="add";
    else if(shares===null && prevShares!==null) op="remove";
    else if(shares!==null && prevShares!==null) {{
      diff = shares - prevShares;
      if(diff>0) op="up"; else if(diff<0) op="down"; else op="flat";
    }}

    const dateLabel = fmtDate(d);
    let badge="", diffHtml="";
    if(op==="add")    {{ badge=`<span class="op-badge op-add">建倉</span>`; }}
    else if(op==="remove") {{ badge=`<span class="op-badge op-remove">清倉</span>`; }}
    else if(op==="up")  {{ badge=`<span class="op-badge op-up">加碼</span>`; diffHtml=`<span class="sh-diff w-up">+${{fmtShares(diff)}}</span>`; }}
    else if(op==="down"){{ badge=`<span class="op-badge op-down">減碼</span>`; diffHtml=`<span class="sh-diff w-down">${{fmtShares(diff)}}</span>`; }}
    else if(op==="flat") {{ badge=`<span class="op-badge" style="background:var(--badge-bg);color:var(--sub)">持平</span>`; }}
    else if(i===0 && shares!==null) {{ badge=`<span class="op-badge" style="background:var(--badge-bg);color:var(--sub)">首筆</span>`; }}

    if(shares===null) {{
      return `<div class="sh-row">
        <span class="sh-date">${{dateLabel}}</span>
        <span></span>
        <span class="sh-none">未持有</span>
      </div>`;
    }}
    return `<div class="sh-row">
      <span class="sh-date">${{dateLabel}}</span>
      ${{badge}}
      <span class="sh-shares">${{fmtShares(shares)}} 股</span>
      <span class="sh-weight">${{weight.toFixed(2)}}%</span>
      ${{diffHtml||"<span></span>"}}
    </div>`;
  }});

  histEl.innerHTML =
    `<div class="sh-header">
      <div><div class="sh-title">${{name}}</div><div class="sh-subtitle">${{code}} ｜ 近 ${{dates.length}} 個交易日</div></div>
      <button class="sh-close" onclick="clearStockHistory('${{etfId}}')">×</button>
    </div>`+rows.join("");
  histEl.style.display = "block";
}}

function clearStockHistory(etfId) {{
  document.getElementById("hist"+etfId).style.display = "none";
  document.getElementById("srch"+etfId).value = "";
}}

["991a","992a","982a","981a","980a"].forEach(id=>{{ initSelects(id); render(id); }});
</script>
</body>
</html>'''


def main():
    dry_run = "--dry-run" in sys.argv
    print("=" * 50)
    print(f"ETF 資料更新  {'(Dry Run)' if dry_run else ''}")
    print("=" * 50)

    if not dry_run:
        download_today()

    print("📂  解析現有資料…")
    fuhua_data, qunyi_data, qunyi982_data, unitrust_data, nomura_data = load_all_data()
    print(f"  復華991：{sorted(fuhua_data.keys())}")
    print(f"  群益992：{sorted(qunyi_data.keys())}")
    print(f"  群益982：{sorted(qunyi982_data.keys())}")
    print(f"  統一981：{sorted(unitrust_data.keys())}")
    print(f"  野村980：{sorted(nomura_data.keys())}")

    print("🖊️   產生 HTML…")
    html = generate_html(fuhua_data, qunyi_data, qunyi982_data, unitrust_data, nomura_data)
    with open(HTML_OUT, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  ✓ 已寫入 {HTML_OUT}")
    print("完成！")


if __name__ == "__main__":
    main()
